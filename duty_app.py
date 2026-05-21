import streamlit as st
import pandas as pd
import io
import json
import os
import xlsxwriter
from pathlib import Path
from openpyxl import load_workbook
from logic import (
    parse_filename_for_name, 
    parse_pdf_schedule, 
    check_availability,
    SHIFTS, DAYS, DAYS_CN,
    CAMPUS_NORTH, CAMPUS_SOUTH,
    ROLE_OFFICER, ROLE_CADRE, ROLE_MINISTER, ROLE_VICE_MINISTER, ROLE_DIRECTOR
)

# --- Configuration & Constants ---
st.set_page_config(page_title="智能值班表排班系统", layout="wide", page_icon="📅")

DUTY_TABLE_COLUMNS = ['Week', 'Day', 'Shift', 'Name', 'Class', 'Department', 'Role', 'Campus']


def get_data_file():
    base_dir = Path.cwd() / "data"
    try:
        base_dir.mkdir(parents=True, exist_ok=True)
    except OSError:
        base_dir = Path.home() / ".DutyScheduler" / "data"
        base_dir.mkdir(parents=True, exist_ok=True)
    return base_dir / "app_state.json"


def serialize_schedules(schedules):
    serializable = {}
    for name, slots in schedules.items():
        serializable[name] = []
        for slot in slots:
            item = dict(slot)
            item["weeks"] = sorted(list(item.get("weeks", [])))
            serializable[name].append(item)
    return serializable


def deserialize_schedules(schedules):
    restored = {}
    for name, slots in schedules.items():
        restored[name] = []
        for slot in slots:
            item = dict(slot)
            item["weeks"] = set(item.get("weeks", []))
            restored[name].append(item)
    return restored


def save_app_state():
    data_file = get_data_file()

    payload = {}
    if "personnel" in st.session_state:
        payload["personnel"] = st.session_state["personnel"].to_dict(orient="records")
    if "schedules" in st.session_state:
        payload["schedules"] = serialize_schedules(st.session_state["schedules"])
    if "duty_table" in st.session_state:
        payload["duty_table"] = st.session_state["duty_table"].to_dict(orient="records")

    data_file.write_text(json.dumps(payload, ensure_ascii=False, indent=2, default=str), encoding="utf-8")


def load_app_state():
    data_file = get_data_file()
    if not data_file.exists():
        return

    try:
        payload = json.loads(data_file.read_text(encoding="utf-8"))
        if "personnel" in payload:
            st.session_state["personnel"] = pd.DataFrame(payload["personnel"])
        if "schedules" in payload:
            st.session_state["schedules"] = deserialize_schedules(payload["schedules"])
        if "duty_table" in payload:
            duty_table = pd.DataFrame(payload["duty_table"])
            for col in DUTY_TABLE_COLUMNS:
                if col not in duty_table.columns:
                    duty_table[col] = pd.Series(dtype="object")
            st.session_state["duty_table"] = duty_table[DUTY_TABLE_COLUMNS]
    except Exception as e:
        st.warning(f"读取本机已保存数据失败：{e}")


def clear_saved_data_file():
    data_file = get_data_file()
    if data_file.exists():
        data_file.unlink()


def get_default_export_dir():
    export_dir = Path.cwd() / "值班表"
    export_dir.mkdir(parents=True, exist_ok=True)
    return export_dir


def sanitize_filename(filename):
    cleaned = str(filename or "").strip()
    for ch in '<>:"/\\|?*':
        cleaned = cleaned.replace(ch, "_")
    cleaned = cleaned.strip(" .")
    if not cleaned:
        cleaned = "值班表.xlsx"
    if not cleaned.lower().endswith(".xlsx"):
        cleaned += ".xlsx"
    return cleaned


def default_duty_export_filename(duty_table):
    weeks = sorted(
        int(week)
        for week in pd.Series(duty_table.get("Week", [])).dropna().unique()
        if str(week).strip() != ""
    )
    if not weeks:
        return "值班表.xlsx"
    if len(weeks) == 1:
        return f"第{weeks[0]}周值班表.xlsx"
    return f"第{weeks[0]}-{weeks[-1]}周值班表.xlsx"


def build_duty_export_df(duty_table):
    return duty_table.rename(columns={
        "Week": "周次",
        "Day": "星期",
        "Shift": "班次",
        "Name": "姓名",
        "Class": "班级",
        "Department": "部门",
        "Role": "职位",
        "Campus": "校区",
    })


def build_duty_excel_bytes(export_df):
    out_buffer = io.BytesIO()
    with pd.ExcelWriter(out_buffer, engine="xlsxwriter") as writer:
        export_df.to_excel(writer, index=False, sheet_name="值班表")
        worksheet = writer.sheets["值班表"]
        worksheet.freeze_panes(1, 0)
        widths = [8, 12, 24, 12, 24, 18, 12, 12]
        for idx, width in enumerate(widths[:len(export_df.columns)]):
            worksheet.set_column(idx, idx, width)
    return out_buffer.getvalue()


def choose_excel_save_path(default_filename):
    import tkinter as tk
    from tkinter import filedialog

    default_dir = get_default_export_dir()
    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)
    try:
        selected = filedialog.asksaveasfilename(
            parent=root,
            title="保存值班表",
            initialdir=str(default_dir),
            initialfile=sanitize_filename(default_filename),
            defaultextension=".xlsx",
            filetypes=[("Excel 工作簿", "*.xlsx"), ("所有文件", "*.*")],
        )
    finally:
        root.destroy()

    if not selected:
        return None
    return Path(selected)


PERSONNEL_COLUMN_MAP = {
    "name": "Name",
    "姓名": "Name",
    "学生姓名": "Name",
    "class": "Class",
    "班级": "Class",
    "专业班级": "Class",
    "专业班": "Class",
    "major": "Major",
    "专业": "Major",
    "grade": "Grade",
    "年级": "Grade",
    "department": "Department",
    "部门": "Department",
    "role": "Role",
    "职位": "Role",
    "职务": "Role",
    "岗位": "Role",
    "角色": "Role",
    "campus": "Campus",
    "校区": "Campus",
    "序号": "Index",
}


def normalize_column_name(value):
    if value is None:
        return ""
    return str(value).replace("\n", "").replace(" ", "").strip()


def canonical_column_name(value):
    key = normalize_column_name(value)
    return PERSONNEL_COLUMN_MAP.get(key, PERSONNEL_COLUMN_MAP.get(key.lower(), key))


def read_excel_with_merged_cells(file_obj):
    wb = load_workbook(file_obj, data_only=True)
    ws = wb.active
    merged_values = {}

    for merged_range in ws.merged_cells.ranges:
        top_value = ws.cell(merged_range.min_row, merged_range.min_col).value
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                merged_values[(row, col)] = top_value

    values = []
    for row in range(1, ws.max_row + 1):
        row_values = []
        for col in range(1, ws.max_column + 1):
            value = ws.cell(row, col).value
            if value is None:
                value = merged_values.get((row, col))
            row_values.append(value)
        values.append(row_values)

    header_idx = None
    for idx, row in enumerate(values):
        mapped = [canonical_column_name(cell) for cell in row]
        hits = sum(col in {"Name", "Class", "Grade", "Department", "Role", "Campus"} for col in mapped)
        if "Name" in mapped and hits >= 2:
            header_idx = idx
            break

    if header_idx is None:
        return pd.DataFrame(values)

    columns = []
    seen = {}
    for idx, cell in enumerate(values[header_idx]):
        col = canonical_column_name(cell) or f"Unnamed_{idx + 1}"
        seen[col] = seen.get(col, 0) + 1
        if seen[col] > 1:
            col = f"{col}_{seen[col]}"
        columns.append(col)

    return pd.DataFrame(values[header_idx + 1:], columns=columns)


def read_personnel_file(file_obj):
    file_name = getattr(file_obj, "name", "").lower()
    if file_name.endswith(".csv"):
        return pd.read_csv(file_obj)
    return read_excel_with_merged_cells(file_obj)


def normalize_campus(value):
    value = str(value).strip()
    if "南" in value:
        return CAMPUS_SOUTH
    if "北" in value:
        return CAMPUS_NORTH
    return value


def normalize_role(row):
    role = str(row.get("Role", "")).strip()
    department = str(row.get("Department", "")).strip()
    role_text = f"{role} {department}"
    if "主任团" in role_text:
        return ROLE_DIRECTOR
    if "副部" in role:
        return ROLE_VICE_MINISTER
    if "部长" in role:
        return ROLE_MINISTER
    if "干部" in role:
        return ROLE_CADRE
    if "干事" in role:
        return ROLE_OFFICER
    return role


def normalize_personnel_dataframe(df_p):
    df_p = df_p.copy()
    df_p = df_p.rename(columns={col: canonical_column_name(col) for col in df_p.columns})

    if "Name" in df_p.columns:
        df_p["Name"] = df_p["Name"].fillna("").astype(str).str.strip()
        df_p = df_p[df_p["Name"] != ""]

    if "Class" not in df_p.columns:
        if "Major" in df_p.columns and "Grade" in df_p.columns:
            df_p["Class"] = (
                df_p["Major"].fillna("").astype(str).str.strip()
                + " "
                + df_p["Grade"].fillna("").astype(str).str.strip()
            ).str.strip()
        elif "Major" in df_p.columns:
            df_p["Class"] = df_p["Major"]
        elif "Grade" in df_p.columns:
            df_p["Class"] = df_p["Grade"]
        else:
            df_p["Class"] = ""

    if "Department" not in df_p.columns:
        df_p["Department"] = ""

    for col in ["Class", "Grade", "Department", "Role", "Campus"]:
        if col in df_p.columns:
            df_p[col] = df_p[col].fillna("").astype(str).str.strip()

    if "Department" in df_p.columns:
        df_p["Department"] = df_p["Department"].replace("", pd.NA).ffill().fillna("")

    if "Campus" in df_p.columns:
        df_p["Campus"] = df_p["Campus"].apply(normalize_campus)
    if "Role" in df_p.columns:
        df_p["Role"] = df_p.apply(normalize_role, axis=1)

    return df_p


def build_personnel_template():
    sample_data = pd.DataFrame({
        "序号": [1, 2, 3, 4],
        "姓名": ["张三", "李四", "王五", "赵六"],
        "部门": ["办公室", "活动部", "宣传部", "主任团"],
        "职务": ["干事", "干事", "干部", "干部"],
        "专业班级": ["会计学2501", "法学2502", "计算机科学与技术2401", "公共事业管理2301"],
        "年级": ["大一", "大一", "大二", "大三"],
        "校区": ["旗山北校区", "旗山南校区", "旗山北校区", "旗山北校区"],
    })
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="xlsxwriter") as writer:
        sample_data.to_excel(writer, index=False, sheet_name="人员名单")
        worksheet = writer.sheets["人员名单"]
        worksheet.freeze_panes(1, 0)
        widths = [8, 12, 18, 12, 26, 10, 16]
        for idx, width in enumerate(widths):
            worksheet.set_column(idx, idx, width)
    return buffer.getvalue()


def unique_filter_options(series):
    return sorted(
        value
        for value in series.fillna("").astype(str).str.strip().unique()
        if value
    )


def person_select_label(name, personnel_df):
    if name == "未安排":
        return name
    matches = personnel_df[personnel_df["Name"] == name]
    if matches.empty:
        return name

    person = matches.iloc[0]
    details = []
    for col in ["Role", "Grade", "Department"]:
        value = str(person.get(col, "")).strip()
        if value:
            details.append(value)

    if not details:
        return name
    return f"{name}（{'，'.join(details)}）"


if "app_state_loaded" not in st.session_state:
    load_app_state()
    st.session_state["app_state_loaded"] = True


def parse_schedule_files(schedule_files):
    parsed = {}
    logs = []
    total = len(schedule_files)
    bar = st.progress(0)
    status_text = st.empty()

    for i, f in enumerate(schedule_files):
        file_name = getattr(f, "name", str(f))
        display_name = Path(file_name).name
        try:
            status_text.text(f"正在处理: {display_name}...")
            name = parse_filename_for_name(display_name)
            slots = parse_pdf_schedule(f)
            parsed[name] = slots
            logs.append(f"✅ {name}: 成功提取 {len(slots)} 个课程时间段")
        except Exception as e:
            logs.append(f"❌ {display_name}: 解析失败 - {e}")
        bar.progress((i + 1) / total)

    status_text.text("处理完成！")
    st.session_state['schedules'] = parsed
    save_app_state()
    st.success(f"🎉 全部处理完成！共成功解析 {len(parsed)} 份课表。")

    if 'personnel' in st.session_state:
        personnel_names = st.session_state['personnel']['Name'].dropna().astype(str)
        all_names = set(personnel_names)
        parsed_names = set(parsed.keys())
        missing = all_names - parsed_names
        missing = [str(m) for m in missing if m and str(m).strip()]

        if missing:
            st.warning(f"⚠️ 注意：以下 {len(missing)} 位同学的人员名单已导入，但未找到对应的课表文件（他们将被视为【不可值班】）：\n" + "、".join(missing))
        else:
            st.success("✅ 完美！所有人都有课表。")

    with st.expander("查看详细处理日志"):
        for l in logs:
            st.write(l)


def find_pdf_files(folder_path, recursive=False):
    if not folder_path or not folder_path.strip():
        raise ValueError("请先填写课表文件夹路径")

    folder = Path(folder_path.strip().strip('"').strip("'")).expanduser()
    if not folder.exists():
        raise FileNotFoundError(f"文件夹不存在：{folder}")
    if not folder.is_dir():
        raise NotADirectoryError(f"这不是文件夹：{folder}")

    pattern = "**/*.pdf" if recursive else "*.pdf"
    return sorted(folder.glob(pattern), key=lambda p: str(p).lower())


# --- Main App ---

st.title("📅 智能部门值班表排班系统")
st.markdown("---")

# Sidebar
st.sidebar.header("⚙️ 全局设置")
target_weeks = st.sidebar.multiselect("选择需要排班的周次", list(range(1, 21)), default=[1])
st.sidebar.info("提示：可以在这里选择多个周次，然后在排班页面切换。")

# Tabs
tab1, tab2, tab3, tab4 = st.tabs(["📂 第一步：数据导入", "🔍 第二步：空闲查询", "📅 第三步：排班制作", "📊 第四步：记录统计"])

# --- Tab 1: Data Upload ---
with tab1:
    st.header("📂 数据导入")
    st.markdown("请按照顺序完成以下两项数据的上传。")
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.subheader("1. 上传人员名单")
        st.markdown("请上传包含人员信息的 Excel 文件。")
        st.info("推荐表头：序号、姓名、部门、职务、专业班级、年级、校区。也兼容英文表头：Name、Class、Department、Role、Campus。")
        
        # Template
        if st.button("📥 下载Excel模板"):
            st.download_button("点击下载模板.xlsx", build_personnel_template(), "人员名单模板.xlsx")

        p_file = st.file_uploader("点击上传人员名单 (Excel/CSV)", type=['xlsx', 'csv'])
        if p_file:
            try:
                df_p = normalize_personnel_dataframe(read_personnel_file(p_file))
                
                # Keep only relevant columns if possible to avoid clutter, or just ensure existence
                # We need Name, Class, Department, Role, Campus
                
                # Validation: Check for required columns
                missing_cols = []
                if 'Name' not in df_p.columns: missing_cols.append('Name')
                if 'Role' not in df_p.columns: missing_cols.append('Role')
                if 'Campus' not in df_p.columns: missing_cols.append('Campus')
                
                if missing_cols:
                    st.error(f"❌ 上传的文件缺少关键列: {', '.join(missing_cols)}。请检查表头是否正确（支持中文：姓名、职务/职位、校区）。")
                else:
                    keep_cols = ['Name', 'Class', 'Grade', 'Department', 'Role', 'Campus']
                    keep_cols = [col for col in keep_cols if col in df_p.columns]
                    df_p = df_p[keep_cols]
                    st.session_state['personnel'] = df_p
                    save_app_state()
                    st.success(f"✅ 成功导入 {len(df_p)} 名人员信息！")
                    departments = unique_filter_options(df_p['Department'])
                    if departments:
                        st.info("已识别部门：" + "、".join(departments))
                    with st.expander("查看导入的人员列表"):
                        st.dataframe(df_p)
            except Exception as e:
                st.error(f"读取文件失败: {e}")

    with col2:
        st.subheader("2. 批量上传课表 (PDF)")
        st.markdown("请选择多份 PDF，或直接导入课表文件夹。")
        st.warning("注意：PDF文件名必须包含学生姓名（例如 `田宇(2023-2024-2)课表.pdf`）")

        import_mode = st.radio(
            "课表导入方式",
            ["选择多个PDF文件", "选择课表文件夹", "填写本机文件夹路径"],
            horizontal=True,
        )

        if import_mode == "选择多个PDF文件":
            pdf_files = st.file_uploader(
                "点击选择PDF课表文件 (支持按住Ctrl/Command键多选)",
                type=['pdf'],
                accept_multiple_files=True,
            )

            if pdf_files:
                st.write(f"已选择 {len(pdf_files)} 个文件")
                if st.button("🚀 开始解析课表", key="parse_pdf_files"):
                    parse_schedule_files(pdf_files)

        elif import_mode == "选择课表文件夹":
            folder_files = st.file_uploader(
                "点击选择课表文件夹",
                type=['pdf'],
                accept_multiple_files="directory",
            )

            if folder_files:
                st.write(f"已从文件夹中选择 {len(folder_files)} 个PDF文件")
                if st.button("🚀 开始解析文件夹课表", key="parse_pdf_directory_upload"):
                    parse_schedule_files(folder_files)

        else:
            folder_path = st.text_input("课表文件夹路径", placeholder=r"例如：D:\课表\2025-2026-1")
            include_subfolders = st.checkbox("同时读取子文件夹中的PDF", value=False)

            if st.button("📁 扫描并解析文件夹", key="parse_pdf_directory_path"):
                try:
                    pdf_paths = find_pdf_files(folder_path, recursive=include_subfolders)
                    if not pdf_paths:
                        st.warning("这个文件夹里没有找到 PDF 文件。")
                    else:
                        st.write(f"已找到 {len(pdf_paths)} 个PDF文件")
                        parse_schedule_files(pdf_paths)
                except Exception as e:
                    st.error(f"读取文件夹失败：{e}")

    st.divider()
    st.subheader("当前数据管理")
    st.caption(f"数据会自动保存在本机：{get_data_file()}")

    personnel_count = len(st.session_state["personnel"]) if "personnel" in st.session_state else 0
    schedule_count = len(st.session_state["schedules"]) if "schedules" in st.session_state else 0
    duty_count = len(st.session_state["duty_table"]) if "duty_table" in st.session_state else 0

    m1, m2, m3 = st.columns(3)
    m1.metric("人员名单", f"{personnel_count} 人")
    m2.metric("课表", f"{schedule_count} 份")
    m3.metric("排班记录", f"{duty_count} 条")

    if schedule_count:
        delete_name = st.selectbox("删除某一份课表", sorted(st.session_state["schedules"].keys()))
        if st.button("🗑️ 删除选中课表"):
            st.session_state["schedules"].pop(delete_name, None)
            save_app_state()
            st.success(f"已删除 {delete_name} 的课表。")
            st.rerun()

    c_clear1, c_clear2, c_clear3, c_clear4 = st.columns(4)
    if c_clear1.button("清空人员名单"):
        st.session_state.pop("personnel", None)
        save_app_state()
        st.rerun()
    if c_clear2.button("清空全部课表"):
        st.session_state.pop("schedules", None)
        save_app_state()
        st.rerun()
    if c_clear3.button("清空排班记录"):
        st.session_state.pop("duty_table", None)
        save_app_state()
        st.rerun()
    if c_clear4.button("清空所有数据"):
        for key in ["personnel", "schedules", "duty_table"]:
            st.session_state.pop(key, None)
        clear_saved_data_file()
        st.rerun()

# --- Tab 2: Search & Filter ---
with tab2:
    st.header("🔍 查询谁有空")
    st.markdown("在这里你可以快速查询某个特定时间段有哪些人是可以值班的。")
    
    if 'personnel' in st.session_state and 'schedules' in st.session_state:
        c1, c2, c3 = st.columns(3)
        with c1:
            s_week = st.number_input("选择周次", 1, 20, 1)
        with c2:
            s_day_cn = st.selectbox("选择星期", DAYS_CN)
            s_day_idx = DAYS_CN.index(s_day_cn)
        with c3:
            # Create a display map for shifts
            shift_display_map = {k: k for k in SHIFTS.keys()} 
            s_shift = st.selectbox("选择班次", list(SHIFTS.keys()))
            
        # Filters
        st.markdown("### 筛选条件")
        f_col1, f_col2, f_col3 = st.columns(3)
        with f_col1:
            f_role = st.multiselect("按职位筛选", unique_filter_options(st.session_state['personnel']['Role']))
        with f_col2:
            f_campus = st.multiselect("按校区筛选", unique_filter_options(st.session_state['personnel']['Campus']))
        with f_col3:
            f_dept = st.multiselect("按部门筛选", unique_filter_options(st.session_state['personnel']['Department']))
            
        if st.button("🔎 开始查询"):
            results = []
            df = st.session_state['personnel']
            
            # Apply Pre-filters
            if f_role: df = df[df['Role'].isin(f_role)]
            if f_campus: df = df[df['Campus'].isin(f_campus)]
            if f_dept: df = df[df['Department'].isin(f_dept)]
            
            for _, person in df.iterrows():
                is_free, code, reason = check_availability(
                    person, s_week, s_day_idx, s_shift, st.session_state['schedules']
                )
                if is_free:
                    results.append({
                        "姓名": person['Name'],
                        "班级": person['Class'],
                        "部门": person['Department'],
                        "职位": person['Role'],
                        "校区": person['Campus'],
                        "状态": "✅ 可值班"
                    })
                # Optional: You could show why they are busy if you wanted
            
            if results:
                st.success(f"共找到 {len(results)} 位符合条件且有空的同学！")
                st.dataframe(pd.DataFrame(results))
            else:
                st.warning("⚠️ 该时间段没有符合条件的空闲人员。")
    else:
        st.info("请先在「数据导入」页面上传人员名单和课表。")

# --- Tab 3: Scheduler ---
with tab3:
    st.header("📅 排班制作")
    
    if 'personnel' in st.session_state and 'schedules' in st.session_state:
        # State for Duty Table
        if 'duty_table' not in st.session_state:
            st.session_state['duty_table'] = pd.DataFrame(columns=DUTY_TABLE_COLUMNS)
            
        # Select Week
        sch_week = st.selectbox("正在制作第几周的班表？", target_weeks, key="sch_week")
        
        st.subheader(f"第 {sch_week} 周值班表预览与编辑")
        st.info("👇 下方表格已根据有课和通勤规则自动过滤人员；姓名后会显示职务、年级和部门，便于人工判断。")
        
        with st.form("scheduler_form"):
            changes = []
            
            for day_i, day_cn in enumerate(DAYS_CN):
                st.markdown(f"### {day_cn}")
                cols = st.columns(3)
                
                for shift_i, (shift_name, conf) in enumerate(SHIFTS.items()):
                    # Find current assignment
                    current = st.session_state['duty_table'][
                        (st.session_state['duty_table']['Week'] == sch_week) &
                        (st.session_state['duty_table']['Day'] == day_cn) &
                        (st.session_state['duty_table']['Shift'] == shift_name)
                    ]
                    current_name = current.iloc[0]['Name'] if not current.empty else "未安排"
                    
                    # Find candidates
                    candidates = ["未安排"]
                    
                    # Optimization: Filter personnel once? No, depends on slot.
                    for _, p in st.session_state['personnel'].iterrows():
                        is_free, _, _ = check_availability(p, sch_week, day_i, shift_name, st.session_state['schedules'])
                        if is_free:
                            candidates.append(p['Name'])
                    
                    # Ensure current is in list
                    if current_name not in candidates:
                        candidates.append(current_name)
                        
                    with cols[shift_i]:
                        # Improve label for readability
                        shift_label = shift_name.split('(')[0] # e.g. "Morning"
                        if "Morning" in shift_name: shift_label = "🌞 上午班 (10:15)"
                        elif "Afternoon 1" in shift_name: shift_label = "⛅ 下午一班 (14:15)"
                        elif "Afternoon 2" in shift_name: shift_label = "🌙 下午二班 (16:00)"
                        
                        sel = st.selectbox(
                            f"{shift_label}", 
                            candidates, 
                            index=candidates.index(current_name),
                            format_func=lambda name: person_select_label(name, st.session_state['personnel']),
                            key=f"sel_{sch_week}_{day_i}_{shift_i}"
                        )
                        changes.append({
                            "Week": sch_week,
                            "Day": day_cn,
                            "Shift": shift_name,
                            "Name": sel
                        })
                st.markdown("---")
            
            if st.form_submit_button("💾 保存当前周排班"):
                # Update State
                # Remove old entries for this week
                base_df = st.session_state['duty_table']
                base_df = base_df[base_df['Week'] != sch_week]
                
                new_rows = []
                for ch in changes:
                    if ch['Name'] != "未安排":
                        # Lookup details
                        p_details = st.session_state['personnel'][
                            st.session_state['personnel']['Name'] == ch['Name']
                        ].iloc[0]
                        new_rows.append({
                            "Week": ch['Week'],
                            "Day": ch['Day'],
                            "Shift": ch['Shift'],
                            "Name": ch['Name'],
                            "Class": p_details['Class'],
                            "Department": p_details['Department'],
                            "Role": p_details['Role'],
                            "Campus": p_details['Campus']
                        })
                
                st.session_state['duty_table'] = pd.concat([base_df, pd.DataFrame(new_rows)], ignore_index=True)
                save_app_state()
                st.success("✅ 保存成功！")
                
        # Export
        if not st.session_state['duty_table'].empty:
            st.divider()
            st.subheader("📤 导出/导入")
            
            col_ex1, col_ex2 = st.columns(2)
            
            with col_ex1:
                st.markdown("**导出Excel文件**")
                export_df = build_duty_export_df(st.session_state['duty_table'])
                excel_bytes = build_duty_excel_bytes(export_df)
                default_export_name = default_duty_export_filename(st.session_state['duty_table'])

                if st.button("💾 选择位置并保存"):
                    try:
                        export_path = choose_excel_save_path(default_export_name)
                        if export_path:
                            export_path.parent.mkdir(parents=True, exist_ok=True)
                            export_path.write_bytes(excel_bytes)
                            st.success(f"已保存到：{export_path}")
                        else:
                            st.info("已取消保存。")
                    except Exception as e:
                        st.error(f"保存失败：{e}")

                st.download_button("📥 浏览器下载值班表.xlsx", excel_bytes, default_export_name)
                
            with col_ex2:
                st.markdown("**导入已有排班进行修改**")
                up_sched = st.file_uploader("上传之前导出的值班表Excel", type=['xlsx'])
                if up_sched:
                    if st.button("📥 合并导入"):
                        try:
                            imp_df = pd.read_excel(up_sched)
                            # Map back if needed or assume format is same (need to be careful with column names)
                            # If we exported with Chinese headers, we need to map back to English for internal state
                            if "周次" in imp_df.columns:
                                imp_df = imp_df.rename(columns={
                                    "周次": "Week", "星期": "Day", "班次": "Shift", 
                                    "姓名": "Name", "班级": "Class", "部门": "Department", "职位": "Role", "校区": "Campus"
                                })
                            
                            st.session_state['duty_table'] = pd.concat([st.session_state['duty_table'], imp_df]).drop_duplicates()
                            save_app_state()
                            st.success("✅ 导入成功！")
                        except Exception as e:
                            st.error(f"导入失败: {e}")

# --- Tab 4: Records ---
with tab4:
    st.header("📊 记录统计")
    
    if 'duty_table' in st.session_state and not st.session_state['duty_table'].empty:
        df = st.session_state['duty_table']
        
        # Count by Person
        counts = df['Name'].value_counts().reset_index()
        counts.columns = ['姓名', '值班次数']
        
        c1, c2 = st.columns([1, 2])
        with c1:
            st.subheader("值班次数排行")
            st.dataframe(counts)
        with c2:
            st.subheader("可视化图表")
            st.bar_chart(counts.set_index('姓名'))
            
        # Detailed Log
        st.subheader("详细值班记录")
        search = st.text_input("🔍 搜索姓名")
        
        display_df = df.rename(columns={
            "Week": "周次", "Day": "星期", "Shift": "班次", 
            "Name": "姓名", "Class": "班级", "Department": "部门", "Role": "职位", "Campus": "校区"
        })
        
        if search:
            st.dataframe(display_df[display_df['姓名'].str.contains(search)])
        else:
            st.dataframe(display_df)
    else:
        st.info("暂无值班记录。请先在「排班制作」页面生成排班。")
